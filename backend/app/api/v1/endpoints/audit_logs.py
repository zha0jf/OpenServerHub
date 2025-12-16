from fastapi import APIRouter, Depends, HTTPException, Query, Request, status
from sqlalchemy.ext.asyncio import AsyncSession
from typing import List, Optional
from datetime import datetime, timedelta
from io import StringIO, BytesIO
import csv
from fastapi.responses import StreamingResponse
import logging

# 检查是否安装了openpyxl库
try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app.core.database import get_async_db
from app.schemas.audit_log import AuditLogListResponse, AuditLog
from app.services.audit_log import AuditLogService
from app.services.auth import get_current_admin_user
from app.models.audit_log import AuditLog as AuditLogModel, AuditAction, AuditResourceType, AuditStatus

router = APIRouter()
logger = logging.getLogger(__name__)

# 更具体的路由定义在前面（优先匹配）

@router.get("/types")
async def get_audit_types(
    current_user = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_async_db),
):
    """
    获取审计操作类型和资源类型
    
    仅管理员用户可以访问。
    """
    logger.debug(f"用户 {current_user.username} 请求审计操作类型和资源类型列表")
    
    # 从AuditAction枚举获取所有操作类型
    action_types = [{"action": action.value} for action in AuditAction]
    
    # 从AuditResourceType枚举获取所有资源类型
    resource_types = [{"resource_type": resource_type.value} for resource_type in AuditResourceType]
    
    result = {
        "action_types": action_types,
        "resource_types": resource_types
    }
    
    logger.info(f"审计类型列表查询完成，用户={current_user.username}，操作类型数={len(action_types)}，资源类型数={len(resource_types)}")
    return result

@router.get("/stats/summary")
async def get_audit_stats_summary(
    days: int = Query(7, ge=1, le=90),
    current_user = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_async_db),
):
    """
    获取审计日志统计摘要
    
    返回过去N天内的操作统计信息。
    
    仅管理员用户可以访问。
    """
    logger.debug(f"用户 {current_user.username} 请求审计日志统计摘要，周期={days}天")
    from sqlalchemy import select, func
    
    start_date = datetime.now() - timedelta(days=days)
    
    # 统计各操作类型的数量
    stmt = select(
        AuditLogModel.action,
        func.count(AuditLogModel.id).label('count')
    ).where(
        AuditLogModel.created_at >= start_date
    ).group_by(
        AuditLogModel.action
    )
    result = await db.execute(stmt)
    action_stats = result.all()
    
    # 统计各操作者的活动
    stmt = select(
        AuditLogModel.operator_username,
        func.count(AuditLogModel.id).label('count')
    ).where(
        AuditLogModel.created_at >= start_date
    ).group_by(
        AuditLogModel.operator_username
    )
    result = await db.execute(stmt)
    operator_stats = result.all()
    
    # 统计失败操作
    stmt = select(func.count(AuditLogModel.id)).where(
        AuditLogModel.created_at >= start_date,
        AuditLogModel.status == 'failed'
    )
    result = await db.execute(stmt)
    failed_count = result.scalar()
    
    # 总操作数
    stmt = select(func.count(AuditLogModel.id)).where(
        AuditLogModel.created_at >= start_date
    )
    result = await db.execute(stmt)
    total_count = result.scalar()
    
    logger.info(f"审计日志统计摘要查询完成，周期={days}天，总操作数={total_count}")
    return {
        "period_days": days,
        "start_date": start_date.isoformat(),
        "end_date": datetime.now().isoformat(),
        "total_operations": total_count,
        "failed_operations": failed_count,
        "success_rate": ((total_count - failed_count) / total_count * 100) if total_count > 0 else 0,
        "actions_breakdown": [
            {
                "action": str(action),
                "count": count
            }
            for action, count in action_stats
        ],
        "top_operators": [
            {
                "username": username,
                "count": count
            }
            for username, count in operator_stats
        ][:10],  # 只返回前10个
    }

@router.get("/export/csv")
async def export_audit_logs_csv(
    skip: int = Query(0, ge=0),
    limit: int = Query(10000, ge=1, le=10000),
    action: str = Query(None),
    operator_id: int = Query(None),
    resource_type: str = Query(None),
    resource_id: int = Query(None),
    start_date: str = Query(None),
    end_date: str = Query(None),
    http_request: Request = None,
    current_user = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_async_db),
):
    """
    导出审计日志为CSV格式
    
    仅管理员用户可以访问。
    """
    logger.debug(f"用户 {current_user.username} 请求导出审计日志为CSV，skip={skip}, limit={limit}")
    
    audit_service = AuditLogService(db)
    
    # 解析日期
    start_datetime = None
    end_datetime = None
    
    try:
        if start_date:
            start_datetime = datetime.fromisoformat(start_date)
        if end_date:
            end_datetime = datetime.fromisoformat(end_date)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无效的日期格式: {str(e)}"
        )
    
    # 转换action为枚举值
    audit_action = None
    if action:
        try:
            audit_action = AuditAction(action.lower())
        except ValueError:
            # 如果action不是有效的枚举值，保持为None（不过滤）
            audit_action = None
    
    logs, total = await audit_service.get_logs(
        skip=skip,
        limit=limit,
        action=audit_action,
        operator_id=operator_id,
        resource_type=resource_type,
        resource_id=resource_id,
        start_date=start_datetime,
        end_date=end_datetime
    )
    
    # 生成CSV
    output = StringIO()
    writer = csv.writer(output)
    
    # 写入表头
    headers = [
        'ID', '操作类型', '状态', '操作者ID', '操作者用户名',
        '资源类型', '资源ID', '资源名称', '操作详情', '操作结果',
        '错误消息', 'IP地址', 'User Agent', '创建时间'
    ]
    writer.writerow(headers)
    
    # 写入数据
    for log in logs:
        writer.writerow([
            log.id,
            str(log.action) if log.action else '',
            log.status,
            log.operator_id,
            log.operator_username or '',
            log.resource_type or '',
            log.resource_id,
            log.resource_name or '',
            log.action_details or '',
            log.result or '',
            log.error_message or '',
            log.ip_address or '',
            log.user_agent or '',
            log.created_at.isoformat() if log.created_at else '',
        ])
    
    # 在返回之前记录导出操作
    audit_service.create_log(
        action=AuditAction.AUDIT_LOG_EXPORT,
        operator_id=current_user.id,
        operator_username=current_user.username,
        resource_type="audit_log",
        action_details={
            "export_format": "csv",
            "skip": skip,
            "limit": limit,
            "filters": {
                "action": action,
                "operator_id": operator_id,
                "resource_type": resource_type,
                "resource_id": resource_id,
                "start_date": start_date,
                "end_date": end_date
            }
        },
        result={"status": "success", "format": "csv"},
        ip_address=http_request.client.host if http_request and http_request.client else "unknown",
        user_agent=http_request.headers.get("user-agent", "unknown") if http_request else "unknown",
        status=AuditStatus.SUCCESS,
    )
    
    # 返回CSV文件流
    output.seek(0)
    csv_content = output.getvalue()
    
    # 使用 UTF-8 with BOM 编码，获得跨平台最佳兼容性
    # Windows Excel 识别 BOM 后能正确打开，Linux/Mac 原生支持 UTF-8
    csv_bytes = csv_content.encode('utf-8-sig')
    
    logger.info(f"审计日志CSV导出完成，用户={current_user.username}，导出记录数={len(logs)}")
    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": "attachment; filename=audit_logs.csv"
        }
    )

@router.get("/export/excel")
async def export_audit_logs_excel(
    skip: int = Query(0, ge=0),
    limit: int = Query(10000, ge=1, le=10000),
    action: str = Query(None),
    operator_id: int = Query(None),
    resource_type: str = Query(None),
    resource_id: int = Query(None),
    start_date: str = Query(None),
    end_date: str = Query(None),
    http_request: Request = None,
    current_user = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_async_db),
):
    """
    导出审计日志为Excel格式
    
    仅管理员用户可以访问。
    """
    logger.debug(f"用户 {current_user.username} 请求导出审计日志为Excel，skip={skip}, limit={limit}")
    if not HAS_OPENPYXL:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel导出功能不可用，请安装openpyxl库"
        )
    
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    
    audit_service = AuditLogService(db)
    
    # 解析日期
    start_datetime = None
    end_datetime = None
    
    try:
        if start_date:
            start_datetime = datetime.fromisoformat(start_date)
        if end_date:
            end_datetime = datetime.fromisoformat(end_date)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无效的日期格式: {str(e)}"
        )
    
    # 转换action为枚举值
    audit_action = None
    if action:
        try:
            audit_action = AuditAction(action.lower())
        except ValueError:
            # 如果action不是有效的枚举值，保持为None（不过滤）
            audit_action = None
    
    logs, total = await audit_service.get_logs(
        skip=skip,
        limit=limit,
        action=audit_action,
        operator_id=operator_id,
        resource_type=resource_type,
        resource_id=resource_id,
        start_date=start_datetime,
        end_date=end_datetime
    )
    
    # 生成Excel
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "审计日志"
    
    # 设置表头
    headers = [
        'ID', '操作类型', '状态', '操作者ID', '操作者用户名',
        '资源类型', '资源ID', '资源名称', '操作详情', '操作结果',
        '错误消息', 'IP地址', 'User Agent', '创建时间'
    ]
    
    # 设置表头样式
    if HAS_OPENPYXL and ws is not None:
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment
    
    # 设置列宽
    if ws is not None:
        ws.column_dimensions['A'].width = 10   # ID
        ws.column_dimensions['B'].width = 20   # 操作类型
        ws.column_dimensions['C'].width = 10   # 状态
        ws.column_dimensions['D'].width = 12   # 操作者ID
        ws.column_dimensions['E'].width = 15   # 操作者用户名
        ws.column_dimensions['F'].width = 15   # 资源类型
        ws.column_dimensions['G'].width = 12   # 资源ID
        ws.column_dimensions['H'].width = 20   # 资源名称
        ws.column_dimensions['I'].width = 25   # 操作详情
        ws.column_dimensions['J'].width = 25   # 操作结果
        ws.column_dimensions['K'].width = 20   # 错误消息
        ws.column_dimensions['L'].width = 15   # IP地址
        ws.column_dimensions['M'].width = 20   # User Agent
        ws.column_dimensions['N'].width = 20   # 创建时间
    
    # 写入数据
    if ws is not None:
        for row_idx, log in enumerate(logs, 2):
            ws.cell(row=row_idx, column=1, value=log.id)
            ws.cell(row=row_idx, column=2, value=str(log.action) if log.action else '')
            ws.cell(row=row_idx, column=3, value=log.status)
            ws.cell(row=row_idx, column=4, value=log.operator_id)
            ws.cell(row=row_idx, column=5, value=log.operator_username or '')
            ws.cell(row=row_idx, column=6, value=log.resource_type or '')
            ws.cell(row=row_idx, column=7, value=log.resource_id)
            ws.cell(row=row_idx, column=8, value=log.resource_name or '')
            ws.cell(row=row_idx, column=9, value=log.action_details or '')
            ws.cell(row=row_idx, column=10, value=log.result or '')
            ws.cell(row=row_idx, column=11, value=log.error_message or '')
            ws.cell(row=row_idx, column=12, value=log.ip_address or '')
            ws.cell(row=row_idx, column=13, value=log.user_agent or '')
            ws.cell(row=row_idx, column=14, value=log.created_at.isoformat() if log.created_at else '')
    
    # 在返回之前记录导出操作
    audit_service.create_log(
        action=AuditAction.AUDIT_LOG_EXPORT,
        operator_id=current_user.id,
        operator_username=current_user.username,
        resource_type="audit_log",
        action_details={
            "export_format": "excel",
            "skip": skip,
            "limit": limit,
            "filters": {
                "action": action,
                "operator_id": operator_id,
                "resource_type": resource_type,
                "resource_id": resource_id,
                "start_date": start_date,
                "end_date": end_date
            }
        },
        result={"status": "success", "format": "excel"},
        ip_address=http_request.client.host if http_request.client else "unknown",
        user_agent=http_request.headers.get("user-agent", "unknown") if http_request else "unknown",
        status=AuditStatus.SUCCESS,
    )
    
    # 保存到BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=audit_logs.xlsx"
        }
    )

@router.post("/cleanup")
async def cleanup_old_audit_logs(
    request_body: dict,
    http_request: Request,
    current_user = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_async_db),
):
    """
    清理过期审计日志
    
    仅管理员用户可以访问。
    """
    logger.debug(f"用户 {current_user.username} 请求清理过期审计日志")
    
    days = request_body.get("days")
    if not days or not isinstance(days, int) or days < 1:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="参数'days'必须是大于0的整数"
        )
    
    # 计算截止日期
    cutoff_date = datetime.now() - timedelta(days=days)
    
    audit_service = AuditLogService(db)
    
    try:
        # 删除过期的日志
        from sqlalchemy import delete
        stmt = delete(AuditLogModel).where(AuditLogModel.created_at < cutoff_date)
        result = await db.execute(stmt)
        delete_count = result.rowcount
        await db.commit()
        
        # 记录清理操作
        await audit_service.create_log(
            action=AuditAction.AUDIT_LOG_CLEANUP,
            operator_id=current_user.id,
            operator_username=current_user.username,
            resource_type="audit_log",
            action_details={
                "days": days,
                "cutoff_date": cutoff_date.isoformat()
            },
            result={
                "deleted_count": delete_count,
                "message": f"成功删除{delete_count}条{cutoff_date.strftime('%Y-%m-%d')}之前的审计日志"
            },
            ip_address=http_request.client.host if http_request.client else "unknown",
            user_agent=http_request.headers.get("user-agent", "unknown"),
            status=AuditStatus.SUCCESS,
        )
        
        return {
            "deleted_count": delete_count,
            "message": f"成功删除{delete_count}条{cutoff_date.strftime('%Y-%m-%d')}之前的审计日志"
        }
    except Exception as e:
        await db.rollback()
        
        # 记录清理失败的操作
        try:
            audit_service = AuditLogService(db)
            await audit_service.create_log(
                action=AuditAction.AUDIT_LOG_CLEANUP,
                operator_id=current_user.id,
                operator_username=current_user.username,
                resource_type="audit_log",
                action_details={
                    "days": days,
                    "cutoff_date": cutoff_date.isoformat()
                },
                result={
                    "deleted_count": 0,
                    "message": f"清理审计日志失败: {str(e)}"
                },
                error_message=str(e),
                ip_address=http_request.client.host if http_request.client else "unknown",
                user_agent=http_request.headers.get("user-agent", "unknown"),
                status=AuditStatus.FAILED,
            )
        except Exception:
            pass  # 忽略记录日志时的错误
        
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"清理审计日志失败: {str(e)}"
        )

@router.get("/{log_id}", response_model=AuditLog)
async def get_audit_log(
    log_id: int,
    current_user = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_async_db),
):
    """
    获取指定ID的审计日志详情
    
    仅管理员用户可以访问。
    """
    
    audit_service = AuditLogService(db)
    log = await audit_service.get_log_by_id(log_id)
    
    if not log:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="审计日志不存在"
        )
    
    return AuditLog.model_validate(log)

@router.get("/", response_model=AuditLogListResponse)
@router.get("", response_model=AuditLogListResponse)
async def get_audit_logs(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    action: str = Query(None),
    operator_id: int = Query(None),
    resource_type: str = Query(None),
    resource_id: int = Query(None),
    start_date: str = Query(None),  # ISO格式日期
    end_date: str = Query(None),    # ISO格式日期
    current_user = Depends(get_current_admin_user),
    db: AsyncSession = Depends(get_async_db),
):
    """
    获取审计日志列表
    
    仅管理员用户可以访问。
    
    查询参数:
    - skip: 跳过的记录数 (默认0)
    - limit: 返回的记录数 (默认100, 最大1000)
    - action: 操作类型过滤 (可选)
    - operator_id: 操作者ID过滤 (可选)
    - resource_type: 资源类型过滤 (可选)
    - resource_id: 资源ID过滤 (可选)
    - start_date: 开始日期 (ISO格式, 如2025-01-01, 可选)
    - end_date: 结束日期 (ISO格式, 如2025-01-31, 可选)
    """
    logger.debug(f"用户 {current_user.username} 请求审计日志列表，skip={skip}, limit={limit}")
    audit_service = AuditLogService(db)
    
    # 解析日期
    start_datetime = None
    end_datetime = None
    
    try:
        if start_date:
            start_datetime = datetime.fromisoformat(start_date)
        if end_date:
            end_datetime = datetime.fromisoformat(end_date)
    except ValueError as e:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"无效的日期格式: {str(e)}"
        )
    
    # 转换action为枚举值
    audit_action = None
    if action:
        try:
            audit_action = AuditAction(action.lower())
        except ValueError:
            # 如果action不是有效的枚举值，保持为None（不过滤）
            audit_action = None
    
    logs, total = await audit_service.get_logs(
        skip=skip,
        limit=limit,
        action=audit_action,
        operator_id=operator_id,
        resource_type=resource_type,
        resource_id=resource_id,
        start_date=start_datetime,
        end_date=end_datetime,
    )
    
    logger.info(f"审计日志列表查询完成，用户={current_user.username}，返回记录数={len(logs)}，总记录数={total}")
    return AuditLogListResponse(
        items=[
            AuditLog.model_validate(log) for log in logs
        ],
        total=total,
        skip=skip,
        limit=limit,
    )